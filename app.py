from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

wb = Workbook()

ws = wb.active
img = Image.open('test.png')
rgb_im = img.convert('RGB')
width, height = img.size
for x in range(width):
    col = get_column_letter(x + 1)
    for y in range(height):
        row = y + 1
        r, g, b = rgb_im.getpixel((x, y))
        hex_col = '{:02x}{:02x}{:02x}'.format(r, g, b)
        fill = PatternFill("solid", fgColor=hex_col)
        ws.column_dimensions[col].width = 3
        ws[col + str(row)].fill = fill

wb.save("sample.xlsx")